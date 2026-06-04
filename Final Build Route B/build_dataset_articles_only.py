"""
build_dataset_articles_only.py
==============================
Articles -> Ruling task (standalone experiment, NOT a controlled comparison).

    INPUT   =  cited statutory article texts only
    TARGET  =  the full ruling headnote (cleaned page_summary)

Uses ALL filtered jinayat rulings (no situation/segmentation needed for this task,
so the fused rulings are kept). Drops only rulings whose citations resolve to zero
article texts. Output: train/val/test JSONL (80/10/10) + manifest.json.

NOTE (carry into write-up): this variant has no case facts in the input, so higher
surface metrics reflect templated headnote structure, not deeper reasoning.

USAGE:
    python build_dataset_articles_only.py --rulings jinayat_segmented.xlsx \
        --articles articles.xlsx --outdir dataset_articles_only [--max-articles 6] [--seed 42]
"""
import argparse, json, random, re, statistics
from collections import Counter
from pathlib import Path


def _norm(s): s=re.sub(r"\s+"," ",(s or "").strip()); return s.replace("أ","ا").replace("إ","ا").replace("آ","ا")
_CANON={"قانون اصول المحاكمات المدنية":"قانون اصول المحاكمات المدنية","قانون السير":"قانون السير الجديد",
"قانون السير الجديد":"قانون السير الجديد","قانون التجارة البرية":"قانون التجارة","قانون التجارة":"قانون التجارة",
"اسلحة وذخائر":"الاسلحة والذخائر","الاسلحة والذخائر":"الاسلحة والذخائر",
"قانون عقود العمل الجماعية والوساطة والتحكيم":"قانون عقود العمل الجماعية","قانون عقود العمل الجماعية":"قانون عقود العمل الجماعية"}
def norm_law(s): s=_norm(s); return _CANON.get(s,s)

def clean_headnote(s):
    s=str(s or ""); s=re.sub(r"^\s*بطاقة الحكم.*?(الاعضاء|الأعضاء)\s+\S+(\s*[-،]\s*\S+)*","",s)
    return s.strip()

def load_article_map(path):
    import openpyxl
    wb=openpyxl.load_workbook(path,read_only=True); ws=wb.active
    H={h:i for i,h in enumerate(c.value for c in next(ws.iter_rows(min_row=1,max_row=1)))}
    amap={}
    for r in ws.iter_rows(min_row=2,values_only=True):
        law=norm_law(r[H["law_name"]]); num=str(r[H["article_number"]]).strip()
        txt=(r[H["arabic_text"]] or "").strip()
        if txt: amap[(law,num)]=txt
    return amap

def load_rulings(path):
    import openpyxl
    wb=openpyxl.load_workbook(path,read_only=True); ws=wb.active
    H={h:i for i,h in enumerate(c.value for c in next(ws.iter_rows(min_row=1,max_row=1)))}
    return [{k:r[i] for k,i in H.items()} for r in ws.iter_rows(min_row=2,values_only=True)]

def parse_cit(cell):
    out=[]
    for ch in (cell or "").split("|"):
        ch=ch.strip()
        if ":" in ch:
            num,law=ch.split(":",1); out.append((num.strip(),norm_law(law)))
    return out

def build_input(articles):
    return "المواد القانونية:\n"+"\n".join(f"المادة {n} ({l}): {t}" for n,l,t in articles)

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--rulings",type=Path,required=True)
    ap.add_argument("--articles",type=Path,required=True)
    ap.add_argument("--outdir",type=Path,default=Path("dataset_articles_only"))
    ap.add_argument("--max-articles",type=int,default=6)
    ap.add_argument("--seed",type=int,default=42)
    args=ap.parse_args(); args.outdir.mkdir(parents=True,exist_ok=True)

    amap=load_article_map(args.articles); print(f"Article texts loaded: {len(amap)}")
    rulings=load_rulings(args.rulings); print(f"Rulings loaded:       {len(rulings)}")

    examples=[]; drop_no_target=drop_no_articles=0; cit_total=cit_matched=0
    for r in rulings:
        target=clean_headnote(r.get("page_summary"))
        if not target or len(target)<10: drop_no_target+=1; continue
        resolved=[]
        for num,law in parse_cit(r.get("old_articles_full")):
            cit_total+=1; txt=amap.get((law,num))
            if txt: cit_matched+=1; resolved.append((num,law,txt))
        if not resolved: drop_no_articles+=1; continue
        resolved=resolved[:args.max_articles]
        examples.append({"ruling_id":r.get("ruling_id"),"input":build_input(resolved),
                         "target":target,"n_articles":len(resolved)})

    print(f"\nUsable examples built: {len(examples)}")
    print(f"  dropped (no target):          {drop_no_target}")
    print(f"  dropped (no article matched): {drop_no_articles}")
    print(f"  citation join coverage:       {cit_matched}/{cit_total} ({100*cit_matched/max(cit_total,1):.1f}%)")

    random.seed(args.seed); random.shuffle(examples)
    n=len(examples); n_tr=int(n*0.8); n_va=int(n*0.1)
    splits={"train":examples[:n_tr],"val":examples[n_tr:n_tr+n_va],"test":examples[n_tr+n_va:]}
    for name,rows in splits.items():
        p=args.outdir/f"{name}.jsonl"
        with open(p,"w",encoding="utf-8") as f:
            for ex in rows: f.write(json.dumps({"input":ex["input"],"target":ex["target"]},ensure_ascii=False)+"\n")
        print(f"  wrote {p}  ({len(rows)} examples)")

    manifest={"task":"articles -> full ruling headnote (standalone, no situation)",
        "note":"no case facts in input; surface metrics reflect templated headnote structure",
        "total_rulings":len(rulings),"usable_examples":n,
        "dropped_no_target":drop_no_target,"dropped_no_article_matched":drop_no_articles,
        "citation_join_total":cit_total,"citation_join_matched":cit_matched,
        "citation_join_pct":round(100*cit_matched/max(cit_total,1),1),
        "split":{k:len(v) for k,v in splits.items()},
        "max_articles_per_example":args.max_articles,"seed":args.seed}
    (args.outdir/"manifest.json").write_text(json.dumps(manifest,ensure_ascii=False,indent=2),encoding="utf-8")
    print(f"  wrote {args.outdir/'manifest.json'}")

    il=[len(e["input"]) for e in examples]; tl=[len(e["target"]) for e in examples]
    print(f"\nInput chars  — median {int(statistics.median(il))}, p90 {sorted(il)[int(0.9*n)]}, max {max(il)}")
    print(f"Target chars — median {int(statistics.median(tl))}, p90 {sorted(tl)[int(0.9*n)]}, max {max(tl)}")

if __name__=="__main__": main()
