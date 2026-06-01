"""
Topic-by-court analysis.

Answers  question: "How are topics related to courts?"

Each court handles different types of disputes, so topics should cluster
by court type. This tool shows the relationship explicitly:
  - For each court, the most common topics
  - For each topic, where it appears across courts
  - A confusion-matrix style overview

USAGE:
    # Default text report
    python topic_court_analysis.py --input data/public/structured/

    # Save a CSV matrix you can open in Excel
    python topic_court_analysis.py --input data/public/structured/ --csv topic_court_matrix.csv

    # Restrict to the top N topics (the long tail of 3000+ tags is noisy)
    python topic_court_analysis.py --input data/public/structured/ --top-topics 50

    # Restrict to the major courts only (drops courts with <20 rulings)
    python topic_court_analysis.py --input data/public/structured/ --min-court-size 20
"""

import argparse
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path


def load_rulings(input_dir: Path):
    rulings = []
    for jf in sorted(input_dir.glob("*.json")):
        try:
            rulings.append(json.loads(jf.read_text(encoding="utf-8")))
        except Exception as e:
            print(f"  Error reading {jf.name}: {e}")
    return rulings


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--csv", type=Path, default=None,
                        help="Also save a CSV matrix (topics x courts)")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Also save a multi-sheet Excel workbook "
                             "(matrix + per-court topics + specificity)")
    parser.add_argument("--top-topics", type=int, default=30,
                        help="Per court, show top N topics (default 30)")
    parser.add_argument("--min-court-size", type=int, default=20,
                        help="Drop courts with fewer than this many rulings")
    parser.add_argument("--top-courts-per-topic", type=int, default=5,
                        help="In the per-topic view, show top N courts (default 5)")
    args = parser.parse_args()

    rulings = load_rulings(args.input)
    print(f"Loaded {len(rulings)} rulings\n")

    # Build the (court, topic) co-occurrence
    # court_topics[court] = Counter({topic: count, ...})
    court_topics = defaultdict(Counter)
    # topic_courts[topic] = Counter({court: count, ...})
    topic_courts = defaultdict(Counter)
    # how many rulings per court (denominator for percentages)
    court_size = Counter()
    # how many rulings per topic (denominator the other way)
    topic_size = Counter()

    for r in rulings:
        court = r.get("court_name") or "(unknown)"
        topics = r.get("topics") or []
        court_size[court] += 1
        for t in topics:
            court_topics[court][t] += 1
            topic_courts[t][court] += 1
            topic_size[t] += 1

    # Filter small courts
    major_courts = [c for c, n in court_size.items() if n >= args.min_court_size]
    major_courts.sort(key=lambda c: -court_size[c])

    # ============================================================
    # VIEW 1: For each court, what topics dominate?
    # ============================================================
    print("=" * 72)
    print("VIEW 1: TOP TOPICS PER COURT")
    print("=" * 72)
    print("Shows: for each court, the most frequent topics and what % of")
    print(f"the court's rulings each topic appears in.")
    print(f"(Courts with <{args.min_court_size} rulings are omitted.)\n")

    for court in major_courts:
        n_rulings = court_size[court]
        print(f"\n{court}  ({n_rulings} rulings)")
        print("─" * 60)
        top = court_topics[court].most_common(args.top_topics)
        # Width formatting for Arabic right-to-left readability
        for topic, count in top:
            pct = 100.0 * count / n_rulings
            bar = "█" * int(pct / 2)  # one block per 2%
            print(f"  {count:5d}  ({pct:5.1f}%)  {bar}  {topic}")

    # ============================================================
    # VIEW 2: For each top topic, which courts use it?
    # ============================================================
    print("\n")
    print("=" * 72)
    print("VIEW 2: TOP TOPICS — WHICH COURTS HANDLE THEM")
    print("=" * 72)
    print("For each of the most common topics across the dataset, the courts")
    print("where it appears most often.\n")

    most_common_topics = [t for t, _ in topic_size.most_common(args.top_topics)]
    for topic in most_common_topics:
        total = topic_size[topic]
        print(f"\n{topic}  ({total} mentions total)")
        print("─" * 60)
        for court, count in topic_courts[topic].most_common(args.top_courts_per_topic):
            court_total = court_size[court]
            pct_of_court = 100.0 * count / court_total
            pct_of_topic = 100.0 * count / total
            print(f"  {count:5d}  {pct_of_topic:5.1f}% of mentions  "
                  f"({pct_of_court:5.1f}% of {court})")

    # ============================================================
    # VIEW 3: Court-specificity score for each topic
    # ============================================================
    # A topic is "court-specific" if it's heavily concentrated in one court.
    # We measure this with: max_court_share = (count in most-frequent court) / (total mentions)
    # 1.0 = topic appears only in one court (highly specific)
    # 1/N_courts = topic is uniformly distributed (not specific at all)
    print("\n")
    print("=" * 72)
    print("VIEW 3: HOW COURT-SPECIFIC IS EACH TOP TOPIC?")
    print("=" * 72)
    print("Specificity = (mentions in top court) / (total mentions).")
    print("1.00 = only in one court.  0.20 with 5 courts = uniform across them.\n")

    specificity_rows = []
    for topic in most_common_topics:
        total = topic_size[topic]
        if total < 10:  # skip very rare topics
            continue
        top_court, top_count = topic_courts[topic].most_common(1)[0]
        specificity = top_count / total
        specificity_rows.append((specificity, topic, total, top_court, top_count))

    # Sort: most court-specific first
    specificity_rows.sort(reverse=True)
    print(f"  {'Specificity':<13}{'Total':<8}{'Top Court':<25}Topic")
    print(f"  {'─' * 13}{'─' * 8}{'─' * 25}{'─' * 30}")
    for spec, topic, total, court, ccount in specificity_rows[:args.top_topics]:
        print(f"  {spec:5.2f}        {total:<8}{court:<25}{topic}")

    # ============================================================
    # Optional CSV export — full topic × court matrix
    # ============================================================
    if args.csv:
        # Rows = topics, Columns = courts, Cells = counts
        all_topics_for_csv = [t for t, _ in topic_size.most_common(args.top_topics)]
        with open(args.csv, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            header = ["topic", "total_mentions"] + major_courts
            writer.writerow(header)
            for topic in all_topics_for_csv:
                row = [topic, topic_size[topic]]
                for court in major_courts:
                    row.append(topic_courts[topic].get(court, 0))
                writer.writerow(row)
        print(f"\n\nCSV matrix saved to: {args.csv}")
        print(f"  Rows: top {len(all_topics_for_csv)} topics")
        print(f"  Cols: {len(major_courts)} courts")

    # ============================================================
    # Optional Excel export — multi-sheet, ready to drop in a Word doc
    # ============================================================
    if args.xlsx:
        try:
            from openpyxl import Workbook
        except ImportError:
            print(f"\n⚠  openpyxl not installed. Run: pip install openpyxl")
            print(f"   Skipping Excel export.")
        else:
            wb = Workbook()

            # Sheet 1: the topic x court count matrix
            ws1 = wb.active
            ws1.title = "Topic x Court Matrix"
            ws1.sheet_view.rightToLeft = True   # Arabic display
            ws1.append(["Topic", "Total Mentions"] + major_courts)
            for topic, _ in topic_size.most_common(args.top_topics):
                row = [topic, topic_size[topic]]
                for court in major_courts:
                    row.append(topic_courts[topic].get(court, 0))
                ws1.append(row)

            # Sheet 2: top topics per court (long format — easier to read)
            ws2 = wb.create_sheet("Top Topics Per Court")
            ws2.sheet_view.rightToLeft = True
            ws2.append(["Court", "Court Rulings", "Topic", "Count", "Percent of Court"])
            for court in major_courts:
                n_rulings = court_size[court]
                for topic, count in court_topics[court].most_common(args.top_topics):
                    pct = round(100.0 * count / n_rulings, 1)
                    ws2.append([court, n_rulings, topic, count, pct])

            # Sheet 3: specificity ranking — most court-specific topics
            ws3 = wb.create_sheet("Court Specificity")
            ws3.sheet_view.rightToLeft = True
            ws3.append(["Topic", "Total Mentions", "Top Court",
                       "Mentions in Top Court", "Specificity"])
            for spec, topic, total, court, ccount in specificity_rows[:args.top_topics]:
                ws3.append([topic, total, court, ccount, round(spec, 2)])

            wb.save(args.xlsx)
            print(f"\nExcel workbook saved to: {args.xlsx}")
            print(f"  Sheet 1: Topic x Court count matrix")
            print(f"  Sheet 2: Top topics per court (long format)")
            print(f"  Sheet 3: Court specificity ranking")

    # ============================================================
    # Brief interpretation guide
    # ============================================================
    print("\n")
    print("=" * 72)
    print("HOW TO READ THIS FOR ASSOCIATION RULES")
    print("=" * 72)
    print("""
  - Topics with HIGH specificity (close to 1.0) are strong signals — they
    are dominated by one court and will produce confident association rules
    like 'topic X => court Y'.

  - Topics with LOW specificity (close to 1/n_courts) appear across many
    courts and are general legal concepts (e.g., 'اثبات' = evidence).
    These will NOT yield useful court rules.

  - For your seq2seq pipeline: the (court, topics) pairing helps the model
    learn which legal vocabulary belongs to which judicial branch — a
    well-calibrated model should rarely generate civil-court reasoning
    when conditioned on a criminal-court case description.
""")


if __name__ == "__main__":
    main()