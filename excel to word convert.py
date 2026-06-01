from openpyxl import load_workbook
from docx import Document
from docx.shared import Inches, Pt
from docx.enum.section import WD_SECTION
from docx.oxml.ns import qn
from docx.oxml import OxmlElement

def excel_to_word_landscape(excel_file, output_word_file):
    """Convert Excel file to Word document in landscape"""
    
    # Load Excel workbook
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Create Word document
    doc = Document()
    
    # Set to landscape
    section = doc.sections[0]
    section.page_height = Inches(8.5)
    section.page_width = Inches(11)
    section.left_margin = Inches(0.5)
    section.right_margin = Inches(0.5)
    
    # Create table with same dimensions as Excel
    rows = ws.max_row
    cols = ws.max_column
    table = doc.add_table(rows=rows, cols=cols)
    table.style = 'Light Grid Accent 1'
    
    # Fill table with Excel data
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        for c_idx, cell in enumerate(row, 1):
            cell_value = str(cell) if cell is not None else ""
            table.cell(r_idx - 1, c_idx - 1).text = cell_value
    
    # Save Word document
    doc.save(output_word_file)
    print(f"✓ Converted: {output_word_file}")

# Usage:
excel_to_word_landscape('topic_comparison_for_review.xlsx', 'output.docx')