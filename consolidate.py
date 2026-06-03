"""
Consolidate scraped JSON into single tidy files, with CORRECTED article/decree
detection baked in (handles singular مادة, plural مواد, dual المادتين, decrees
مرسوم, slash-chains, dashes, parentheses, paragraph refs, and Arabic-Indic
digits — the forms the original extractor missed).

USAGE
    # Rulings -> one CSV + a per-court statistics report
    python consolidate.py --rulings data/public/structured_combined --outdir out

    # Article texts -> one CSV (for later joining to rulings)
    python consolidate.py --articles data/articles/structured --outdir out

    # Both, and also write .xlsx versions
    python consolidate.py --rulings R --articles A --outdir out --xlsx

OUTPUTS (in --outdir)
    rulings.csv            one row per ruling, with corrected flags
    rulings_stats.csv      per-court: total / >=1 article / >=1 decree / >=1 ref / zero
    articles.csv           one row per legal article text
(Add --xlsx to also get rulings.xlsx, rulings_stats.xlsx, articles.xlsx — RTL, formatted.)

WHY page_summary, not full_text: every ruling's full_text ends with a site
footer that cites the center's founding decrees (مرسوم 3144/4166/4141). Scanning
full_text would mark every ruling as citing a decree. The summary is clean.
"""

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

# ----------------------------------------------------------------------------
# Reference detector  (validated on 32 positive/negative cases + real rulings)
# ----------------------------------------------------------------------------
_DIG = r'[0-9\u0660-\u0669]'                       # ASCII + Arabic-Indic digits
_RE_ART_DUAL = re.compile(r'ماد(?:تين|تان|تي|تا)')  # المادتين / المادتان ...
_RE_ART_STEM = re.compile(r'ماد[ةته]')             # مادة المادة مادته (ة/ه/ت)
_RE_PLURAL   = re.compile(r'مواد')                 # مواد المواد للمواد بالمواد
_RE_DECREE   = re.compile(r'مرسوم|مراسيم')          # مرسوم المرسوم مراسيم
_RE_LAWNUM   = re.compile(r'قانون|اتفاقية|معاهدة|نظام')

# A number must follow the keyword, separated only by connectors (رقم/عدد,
# punctuation, spaces). This is what excludes "مادة ضرب" (matter), "المواد
# الاولية" (raw materials), "المرسوم المذكور" (numberless), etc.
_ART_TAIL = re.compile(r'\s*(?:رقم|عدد)?\s*[/()\-]*\s*' + _DIG)
_DEC_TAIL = re.compile(r'\s*(?:الاشتراعي|التطبيقي|التنظيمي)?\s*(?:رقم|عدد)?\s*[/()\-]*\s*' + _DIG)
_LAW_TAIL = re.compile(r'\s*(?:[\u0600-\u06FF]+\s+){0,3}(?:رقم|عدد)\s*[/()\-]*\s*' + _DIG)


def detect_references(text: str) -> dict:
    """Return booleans for article / decree / law-or-convention number refs."""
    art = dec = law = False
    if not text:
        return {"article": False, "decree": False, "law_num": False}
    for rx in (_RE_ART_DUAL, _RE_ART_STEM, _RE_PLURAL):
        for m in rx.finditer(text):
            if _ART_TAIL.match(text[m.end():m.end() + 15]):
                art = True
                break
        if art:
            break
    for m in _RE_DECREE.finditer(text):
        if _DEC_TAIL.match(text[m.end():m.end() + 30]):
            dec = True
            break
    for m in _RE_LAWNUM.finditer(text):
        if _LAW_TAIL.match(text[m.end():m.end() + 40]):
            law = True
            break
    return {"article": art, "decree": dec, "law_num": law}


# ----------------------------------------------------------------------------
# Loaders
# ----------------------------------------------------------------------------
def _load_json_dir(path: Path):
    files = sorted(path.glob("*.json"))
    for f in files:
        try:
            yield f, json.loads(f.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  ! skip {f.name}: {e}")


def build_ruling_rows(rulings_dir: Path):
    rows = []
    for _, d in _load_json_dir(rulings_dir):
        summary = d.get("page_summary") or ""
        ref = detect_references(summary)
        old_full = d.get("cited_articles_full") or []
        old_n = len(old_full)

        # Union: old extractor captured articles (incl. sidebar). Text scan adds
        # missed articles AND decrees (the old method never looked for decrees).
        has_article = (old_n > 0) or ref["article"]
        has_decree = ref["decree"]
        has_reference = has_article or has_decree

        rows.append({
            "ruling_id": d.get("ruling_id"),
            "court_name": d.get("court_name"),
            "ruling_date": d.get("ruling_date"),
            "ruling_year": d.get("ruling_year"),
            "is_complete": d.get("is_complete"),
            "has_french": d.get("has_french"),
            "old_n_articles": old_n,
            "old_articles": ";".join(d.get("cited_articles") or []),
            "old_articles_full": " | ".join(
                f"{a.get('number')}:{a.get('law')}" for a in old_full),
            "text_has_article": ref["article"],
            "text_has_decree": ref["decree"],
            "text_has_law_num": ref["law_num"],
            "has_article": has_article,
            "has_decree": has_decree,
            "has_reference": has_reference,
            "classification": "has_reference" if has_reference else "zero",
            "n_topics": len(d.get("topics") or []),
            "topics": " | ".join(d.get("topics") or []),
            "page_summary": summary,
            "source_url": d.get("source_url"),
        })
    return rows


def build_article_rows(articles_dir: Path):
    rows = []
    for _, d in _load_json_dir(articles_dir):
        txt = d.get("arabic_text") or ""
        rows.append({
            "law_name": d.get("law_name"),
            "law_id": d.get("law_id"),
            "article_number": d.get("article_number"),
            "article_id": d.get("article_id"),
            "section_id": d.get("section_id"),
            "is_amended": d.get("is_amended"),
            "text_length": len(txt),
            "arabic_text": txt,
            "source_url": d.get("source_url"),
        })
    return rows


# ----------------------------------------------------------------------------
# Per-court statistics
# ----------------------------------------------------------------------------
def build_stats(ruling_rows):
    by_court = defaultdict(lambda: dict(total=0, art=0, dec=0, ref=0,
                                        old_has=0, recovered=0))
    for r in ruling_rows:
        c = by_court[r["court_name"] or "(unknown)"]
        c["total"] += 1
        c["art"] += int(r["has_article"])
        c["dec"] += int(r["has_decree"])
        c["ref"] += int(r["has_reference"])
        old_has = r["old_n_articles"] > 0
        c["old_has"] += int(old_has)
        if (not old_has) and r["has_reference"]:
            c["recovered"] += 1

    stats = []
    for court, c in sorted(by_court.items(), key=lambda kv: -kv[1]["total"]):
        stats.append({
            "court": court,
            "total_rulings": c["total"],
            "with_>=1_article": c["art"],
            "with_>=1_decree": c["dec"],
            "with_>=1_reference": c["ref"],
            "zero_references": c["total"] - c["ref"],
            "old_method_with_articles": c["old_has"],
            "old_method_zero": c["total"] - c["old_has"],
            "recovered_by_fix": c["recovered"],
        })
    return stats


# ----------------------------------------------------------------------------
# Writers
# ----------------------------------------------------------------------------
def write_csv(rows, path: Path):
    import csv
    if not rows:
        print(f"  (no rows for {path.name})")
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote {path}  ({len(rows)} rows)")


def write_xlsx(rows, path: Path, sheet="Sheet1", total_cols=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    if not rows:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.sheet_view.rightToLeft = True
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="DDDDDD", end_color="DDDDDD")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in rows:
        ws.append([r[h] for h in headers])
    # Optional SUM totals row for numeric stat columns
    if total_cols:
        n = len(rows)
        total_row = []
        for i, h in enumerate(headers, 1):
            if h in total_cols:
                col = ws.cell(row=1, column=i).column_letter
                total_row.append(f"=SUM({col}2:{col}{n + 1})")
            else:
                total_row.append("TOTAL" if i == 1 else "")
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
    print(f"  wrote {path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rulings", type=Path, help="Dir of ruling JSON files")
    ap.add_argument("--articles", type=Path, help="Dir of article JSON files")
    ap.add_argument("--outdir", type=Path, default=Path("out"))
    ap.add_argument("--xlsx", action="store_true", help="Also write .xlsx")
    args = ap.parse_args()
    args.outdir.mkdir(parents=True, exist_ok=True)

    if args.rulings:
        print(f"Rulings <- {args.rulings}")
        rows = build_ruling_rows(args.rulings)
        write_csv(rows, args.outdir / "rulings.csv")
        stats = build_stats(rows)
        write_csv(stats, args.outdir / "rulings_stats.csv")
        if args.xlsx:
            write_xlsx(rows, args.outdir / "rulings.xlsx", "Rulings")
            write_xlsx(stats, args.outdir / "rulings_stats.xlsx", "Stats",
                       total_cols={"total_rulings", "with_>=1_article",
                                   "with_>=1_decree", "with_>=1_reference",
                                   "zero_references", "old_method_with_articles",
                                   "old_method_zero", "recovered_by_fix"})
        print("\n=== PER-COURT STATISTICS ===")
        for s in stats:
            print(f"\n{s['court']}  (total {s['total_rulings']})")
            print(f"   >=1 article:   {s['with_>=1_article']}")
            print(f"   >=1 decree:    {s['with_>=1_decree']}")
            print(f"   >=1 reference: {s['with_>=1_reference']}")
            print(f"   zero:          {s['zero_references']}")
            print(f"   (old method said {s['old_method_zero']} zero; "
                  f"fix recovered {s['recovered_by_fix']})")

    if args.articles:
        print(f"\nArticles <- {args.articles}")
        arows = build_article_rows(args.articles)
        write_csv(arows, args.outdir / "articles.csv")
        if args.xlsx:
            write_xlsx(arows, args.outdir / "articles.xlsx", "Articles")


if __name__ == "__main__":
    main()