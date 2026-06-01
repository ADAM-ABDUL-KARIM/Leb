import json, glob, csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Collect rulings with no cited articles
rulings = []
for f in glob.glob('data/public/structured_combined/*.json'):
    d = json.load(open(f, encoding='utf-8'))
    if not d.get('cited_articles_full'):
        rulings.append(d)

print(f'Total rulings with 0 articles: {len(rulings)}')

# Print samples
for r in rulings[:3]:
    # Corrected line using single quotes for the f-string to prevent syntax errors
    print(f'\n--- {r.get("ruling_id")} | {r.get("court_name")} | {r.get("ruling_date")} ---')
    print(f'TOPICS: {r.get("topics")}')
    print(f'SUMMARY: {str(r.get("page_summary",""))[:250]}...')

# Export to CSV
with open('zero_article_rulings.csv', 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(['ruling_id','court','date','topics','summary'])
    for r in rulings:
        w.writerow([r.get('ruling_id',''), r.get('court_name',''), r.get('ruling_date',''), '|'.join(r.get('topics') or []), r.get('page_summary','')])

# Export to Excel
wb = Workbook()
ws = wb.active
ws.title = 'Zero Article Rulings'
ws.sheet_view.rightToLeft = True

# Add header
ws.append(['ruling_id','court','date','topics','summary'])
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

# Add data rows
for r in rulings:
    ws.append([r.get('ruling_id',''), r.get('court_name',''), r.get('ruling_date',''), '|'.join(r.get('topics') or []), r.get('page_summary','')])

# Formatting
for row in ws.iter_rows(min_row=2):
    for cell in row: 
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 100

wb.save('zero_article_rulings.xlsx')
print(f'Saved: zero_article_rulings.csv + .xlsx ({len(rulings)} rows)')