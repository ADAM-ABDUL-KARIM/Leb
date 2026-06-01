"""
Side-by-side topic comparison for  review.

For every topic that appears in BOTH an article-less ruling AND an article-having
ruling, output one paired example so she can compare and decide whether to keep
or drop the article-less rulings.

USAGE:
    python topic_comparison.py --input data/public/structured_combined/ \
                               --xlsx topic_comparison.xlsx

OUTPUT:
    Excel workbook with three sheets:
    1. Comparison    — one row per topic: article-less ruling vs. with-article ruling
    2. No-Article Only  — topics that appear ONLY in article-less rulings (no comparison possible)
    3. Summary stats — counts, totals, completeness
"""

import argparse
import json
import random
from collections import defaultdict
from pathlib import Path


def load_rulings(input_dir: Path):
    rulings = []
    for jf in sorted(input_dir.glob("*.json")):
        try:
            rulings.append(json.loads(jf.read_text(encoding="utf-8")))
        except Exception as e:
            print(f"  Error reading {jf.name}: {e}")
    return rulings


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True,
                        help="Folder of JSON rulings (e.g., the combined trial-court folder)")
    parser.add_argument("--xlsx", type=Path, required=True,
                        help="Output Excel filename")
    parser.add_argument("--seed", type=int, default=42,
                        help="Random seed for sample selection (reproducible)")
    args = parser.parse_args()

    random.seed(args.seed)

    rulings = load_rulings(args.input)
    print(f"Loaded {len(rulings)} rulings")

    # Split rulings into two pools
    with_articles = [r for r in rulings if r.get("cited_articles")]
    no_articles = [r for r in rulings if not r.get("cited_articles")]
    print(f"  With articles:    {len(with_articles)}")
    print(f"  Without articles: {len(no_articles)}")

    # Build topic -> list of rulings, separated by pool
    topic_no_article = defaultdict(list)   # topic -> [rulings with this topic AND no articles]
    topic_with_article = defaultdict(list) # topic -> [rulings with this topic AND articles]

    for r in no_articles:
        for t in r.get("topics") or []:
            topic_no_article[t].append(r)

    for r in with_articles:
        for t in r.get("topics") or []:
            topic_with_article[t].append(r)

    # For each topic in the no-article pool, find a matching with-article example
    # Sort topics by how many no-article rulings they appear in (most common first)
    comparison_rows = []
    no_article_only_rows = []

    topics_in_no_article = sorted(
        topic_no_article.keys(),
        key=lambda t: -len(topic_no_article[t]),
    )

    for topic in topics_in_no_article:
        # Pick one no-article example for this topic (longest summary = most informative)
        no_art_pool = topic_no_article[topic]
        no_art_example = max(no_art_pool, key=lambda r: len(r.get("page_summary") or ""))

        if topic in topic_with_article:
            # Pick the article-having example with longest summary too
            with_art_pool = topic_with_article[topic]
            with_art_example = max(with_art_pool, key=lambda r: len(r.get("page_summary") or ""))

            comparison_rows.append({
                "topic": topic,
                "no_article_total": len(no_art_pool),
                "with_article_total": len(with_art_pool),
                "no_article_id": no_art_example.get("ruling_id"),
                "no_article_court": no_art_example.get("court_name"),
                "no_article_date": no_art_example.get("ruling_date"),
                "no_article_summary": no_art_example.get("page_summary") or "",
                "with_article_id": with_art_example.get("ruling_id"),
                "with_article_court": with_art_example.get("court_name"),
                "with_article_date": with_art_example.get("ruling_date"),
                "with_article_summary": with_art_example.get("page_summary") or "",
                "with_article_articles": "; ".join(
                    f"المادة {a.get('number','?')} - {a.get('law','?')}"
                    for a in (with_art_example.get("cited_articles_full") or [])
                ),
            })
        else:
            # Topic appears ONLY in no-article rulings — no comparison possible
            no_article_only_rows.append({
                "topic": topic,
                "count": len(no_art_pool),
                "example_id": no_art_example.get("ruling_id"),
                "example_court": no_art_example.get("court_name"),
                "example_date": no_art_example.get("ruling_date"),
                "example_summary": no_art_example.get("page_summary") or "",
            })

    print(f"\n  Topics with comparison available: {len(comparison_rows)}")
    print(f"  Topics ONLY in no-article rulings: {len(no_article_only_rows)}")

    # === Write Excel ===
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise SystemExit("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()

    # ---- Sheet 1: Side-by-side comparison ----
    ws1 = wb.active
    ws1.title = "Comparison"
    ws1.sheet_view.rightToLeft = True

    headers = [
        "Topic",
        "# rulings (no articles)",
        "# rulings (with articles)",
        "—",
        "No-Article Example: ID",
        "No-Article Example: Court",
        "No-Article Example: Date",
        "No-Article Example: Summary",
        "—",
        "With-Article Example: ID",
        "With-Article Example: Court",
        "With-Article Example: Date",
        "With-Article Example: Summary",
        "With-Article Example: Articles Cited",
    ]
    ws1.append(headers)

    # Bold header
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in comparison_rows:
        ws1.append([
            row["topic"],
            row["no_article_total"],
            row["with_article_total"],
            "",
            row["no_article_id"],
            row["no_article_court"],
            row["no_article_date"],
            row["no_article_summary"],
            "",
            row["with_article_id"],
            row["with_article_court"],
            row["with_article_date"],
            row["with_article_summary"],
            row["with_article_articles"],
        ])

    # Column widths
    widths = [25, 12, 12, 3, 12, 20, 12, 80, 3, 12, 20, 12, 80, 40]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # Wrap text in summary columns
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Sheet 2: Topics ONLY in no-article rulings ----
    ws2 = wb.create_sheet("No-Article Only Topics")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["Topic", "# Rulings", "Example ID", "Court", "Date", "Example Summary"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for row in no_article_only_rows:
        ws2.append([
            row["topic"],
            row["count"],
            row["example_id"],
            row["example_court"],
            row["example_date"],
            row["example_summary"],
        ])

    for i, w in enumerate([25, 12, 12, 20, 12, 80], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Sheet 3: Summary ----
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_view.rightToLeft = True
    ws3.append(["Metric", "Value"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    summary_data = [
        ("Total rulings in folder", len(rulings)),
        ("Rulings with cited articles", len(with_articles)),
        ("Rulings without cited articles", len(no_articles)),
        ("", ""),
        ("Topics appearing in article-less rulings", len(topic_no_article)),
        ("Topics with both pools (comparable)", len(comparison_rows)),
        ("Topics ONLY in article-less rulings", len(no_article_only_rows)),
    ]
    for k, v in summary_data:
        ws3.append([k, v])

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 15

    wb.save(args.xlsx)
    print(f"\n  Saved: {args.xlsx}")
    print(f"\n  Sheet 1 (Comparison)        — {len(comparison_rows)} rows")
    print(f"  Sheet 2 (No-Article Only)   — {len(no_article_only_rows)} rows")
    print(f"  Sheet 3 (Summary stats)")


if __name__ == "__main__":
    main()