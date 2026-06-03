"""
build_dataset.py
================
Build the mT5/AraT5 fine-tuning dataset for the task:

    INPUT   =  cited article texts  +  the case situation
    TARGET  =  the court's reasoning + verdict

Only rulings with a usable situation span are kept (segment_status in
{clean, partial} AND non-empty situation). Rulings whose citations resolve to
zero article texts are dropped. Output: train/val/test JSONL (80/10/10),
each line {"input": ..., "target": ...}, plus manifest.json for the write-up.

USAGE:
    python build_dataset.py --rulings jinayat_segmented.xlsx \
        --articles articles.xlsx --outdir dataset [--max-articles 6] [--seed 42]
"""

import argparse, json, random, re, statistics
from collections import Counter
from pathlib import Path


def _norm(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    return s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")


_CANON = {
    "قانون اصول المحاكمات المدنية": "قانون اصول المحاكمات المدنية",
    "قانون السير": "قانون السير الجديد",
    "قانون السير الجديد": "قانون السير الجديد",
    "قانون التجارة البرية": "قانون التجارة",
    "قانون التجارة": "قانون التجارة",
    "اسلحة وذخائر": "الاسلحة والذخائر",
    "الاسلحة والذخائر": "الاسلحة والذخائر",
    "قانون عقود العمل الجماعية والوساطة والتحكيم": "قانون عقود العمل الجماعية",
    "قانون عقود العمل الجماعية": "قانون عقود العمل الجماعية",
}


def norm_law(s: str) -> str:
    s = _norm(s)
    return _CANON.get(s, s)


def load_article_map(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True); ws = wb.active
    H = {h: i for i, h in enumerate(c.value for c in next(ws.iter_rows(min_row=1, max_row=1)))}
    amap = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        law = norm_law(r[H["law_name"]]); num = str(r[H["article_number"]]).strip()
        txt = (r[H["arabic_text"]] or "").strip()
        if txt:
            amap[(law, num)] = txt
    return amap


def load_rulings(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True); ws = wb.active
    H = {h: i for i, h in enumerate(c.value for c in next(ws.iter_rows(min_row=1, max_row=1)))}
    return [{k: r[i] for k, i in H.items()} for r in ws.iter_rows(min_row=2, values_only=True)]


def parse_citations(cell: str):
    out = []
    for chunk in (cell or "").split("|"):
        chunk = chunk.strip()
        if ":" in chunk:
            num, law = chunk.split(":", 1)
            out.append((num.strip(), norm_law(law)))
    return out


def build_input(situation: str, articles: list) -> str:
    parts = [f"المادة {num} ({law}): {txt}" for num, law, txt in articles]
    return "المواد القانونية:\n" + "\n".join(parts) + "\n\nالوقائع:\n" + situation.strip()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rulings", type=Path, required=True)
    ap.add_argument("--articles", type=Path, required=True)
    ap.add_argument("--outdir", type=Path, default=Path("dataset"))
    ap.add_argument("--max-articles", type=int, default=6)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()
    args.outdir.mkdir(parents=True, exist_ok=True)

    amap = load_article_map(args.articles)
    print(f"Article texts loaded: {len(amap)}")
    rulings = load_rulings(args.rulings)
    print(f"Rulings loaded:       {len(rulings)}")
    status_counts = Counter(r.get("segment_status") for r in rulings)
    print(f"Segment status:       {dict(status_counts)}")

    examples = []
    drop_no_situation = drop_no_articles = drop_no_target = 0
    cit_total = cit_matched = 0

    for r in rulings:
        status = r.get("segment_status")
        situation = (r.get("situation") or "").strip()
        target = (r.get("reasoning_verdict") or "").strip()
        if status not in ("clean", "partial") or not situation:
            drop_no_situation += 1; continue
        if not target:
            drop_no_target += 1; continue
        resolved = []
        for num, law in parse_citations(r.get("old_articles_full")):
            cit_total += 1
            txt = amap.get((law, num))
            if txt:
                cit_matched += 1; resolved.append((num, law, txt))
        if not resolved:
            drop_no_articles += 1; continue
        resolved = resolved[: args.max_articles]
        examples.append({"ruling_id": r.get("ruling_id"),
                         "input": build_input(situation, resolved),
                         "target": target, "segment_status": status,
                         "n_articles": len(resolved)})

    print(f"\nUsable examples built: {len(examples)}")
    print(f"  dropped (no situation/fused): {drop_no_situation}")
    print(f"  dropped (no article matched): {drop_no_articles}")
    print(f"  dropped (no target text):     {drop_no_target}")
    print(f"  citation join coverage:       {cit_matched}/{cit_total} "
          f"({100*cit_matched/max(cit_total,1):.1f}%)")

    random.seed(args.seed); random.shuffle(examples)
    n = len(examples); n_tr = int(n*0.8); n_va = int(n*0.1)
    splits = {"train": examples[:n_tr], "val": examples[n_tr:n_tr+n_va],
              "test": examples[n_tr+n_va:]}
    for name, rows in splits.items():
        p = args.outdir / f"{name}.jsonl"
        with open(p, "w", encoding="utf-8") as f:
            for ex in rows:
                f.write(json.dumps({"input": ex["input"], "target": ex["target"]},
                                   ensure_ascii=False) + "\n")
        print(f"  wrote {p}  ({len(rows)} examples)")

    manifest = {
        "task": "(articles + situation) -> (reasoning + verdict)",
        "model_target": "AraT5v2-base / mT5",
        "total_rulings_in_file": len(rulings),
        "segment_status_counts": dict(status_counts),
        "usable_examples": n,
        "dropped_no_situation_or_fused": drop_no_situation,
        "dropped_no_article_matched": drop_no_articles,
        "dropped_no_target": drop_no_target,
        "citation_join_total": cit_total, "citation_join_matched": cit_matched,
        "citation_join_pct": round(100*cit_matched/max(cit_total,1), 1),
        "split": {k: len(v) for k, v in splits.items()},
        "max_articles_per_example": args.max_articles, "seed": args.seed,
    }
    (args.outdir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  wrote {args.outdir / 'manifest.json'}")

    in_lens = [len(e["input"]) for e in examples]
    tg_lens = [len(e["target"]) for e in examples]
    print(f"\nInput chars  — median {int(statistics.median(in_lens))}, "
          f"p90 {sorted(in_lens)[int(0.9*n)]}, max {max(in_lens)}")
    print(f"Target chars — median {int(statistics.median(tg_lens))}, "
          f"p90 {sorted(tg_lens)[int(0.9*n)]}, max {max(tg_lens)}")


if __name__ == "__main__":
    main()
