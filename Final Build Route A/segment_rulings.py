"""
segment_rulings.py
==================
Split each jinayat ruling HEADNOTE into two spans using the Anthropic API:
  - situation        : the factual narrative (what happened)
  - reasoning_verdict: the court's legal reasoning, qualification, and outcome

WHY: the dataset task is  (articles + situation) -> (reasoning + verdict).
The scraped page_summary fuses both, so we segment it with an LLM. The split is
imperfect by nature (spans bleed, ~some headnotes lead with the verdict); the
script records a quality flag per row so the imperfection is measurable, not hidden.

INPUT  : the filtered jinayat Excel (we re-apply the filter here so the script is
         self-contained: is_complete AND has_article AND not boilerplate).
OUTPUT : jinayat_segmented.xlsx  (all original columns + situation,
         reasoning_verdict, segment_status), written incrementally so a crash or
         stop never loses finished work and never re-charges for done rows.

SAFETY:
  * --limit N  : process only the first N rulings (use --limit 10 to test cost
                 and quality BEFORE the full run).
  * checkpoint : results are saved to a JSONL cache after every row; re-running
                 skips rulings already done. You never pay twice for a ruling.
  * the script prints a running token + cost estimate so you watch spend live.

USAGE:
  # 0. one-time: create an API key at console.anthropic.com, load $5 of credit
  # 1. set the key (PowerShell):   $env:ANTHROPIC_API_KEY = "sk-ant-..."
  #    (or pass --api-key on the command line)
  # 2. TEST first (10 rulings):
  #        python segment_rulings.py --input All_jinayat_rulings.xlsx --limit 10
  #    inspect jinayat_segmented.xlsx, confirm spans look right + check printed cost
  # 3. FULL run (remove --limit):
  #        python segment_rulings.py --input All_jinayat_rulings.xlsx
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

# ---- pricing (USD per million tokens) for the model we use -----------------
# Sonnet-class pricing; used only to PRINT an estimate so you can watch spend.
# Adjust if you switch models. These are order-of-magnitude guides, not billing.
PRICE_IN_PER_MTOK = 3.00
PRICE_OUT_PER_MTOK = 15.00
MODEL = "claude-sonnet-4-6"   # strong Arabic; swap to a haiku model to cut cost

# ---- the segmentation prompt ------------------------------------------------
# Carefully scoped: Arabic legal headnote in, strict JSON out. We instruct the
# model to copy spans verbatim (no paraphrase), and to handle the verdict-first
# / fused cases by marking them rather than inventing a situation.
SYSTEM_PROMPT = """أنت مساعد متخصص في تحليل النصوص القانونية اللبنانية. مهمتك تقسيم ملخص حكم جزائي إلى قسمين دون إعادة صياغة، بالنقل الحرفي فقط:

1. "situation": السرد الوقائعي — ما الذي حدث فعلاً (الأفعال المنسوبة، الظروف، ما ضُبط، أقوال الأطراف). الوقائع المجردة دون التكييف القانوني أو الحكم.

2. "reasoning_verdict": تعليل المحكمة وتكييفها القانوني ونتيجتها (الإدانة/التبرئة/إعادة التكييف، المواد المطبَّقة، الأسباب المخففة، العقوبة).

قواعد صارمة:
- انقل النص حرفياً من الملخص. لا تُعِد الصياغة ولا تختصر ولا تضف.
- إذا كان الحكم يبدأ بالنتيجة أو يدمج الوقائع والتعليل في جملة واحدة بحيث يتعذّر فصل سرد وقائعي مستقل، اترك "situation" فارغاً ("") وضع النص كاملاً في "reasoning_verdict"، واضبط "status" على "fused".
- إذا أمكن الفصل بوضوح اضبط "status" على "clean".
- إذا أمكن الفصل جزئياً مع تداخل اضبط "status" على "partial".

أعد فقط JSON بالشكل التالي دون أي نص آخر:
{"situation": "...", "reasoning_verdict": "...", "status": "clean|partial|fused"}"""


def clean_boiler(s: str) -> str:
    s = str(s or "")
    s = re.sub(r"^\s*بطاقة الحكم.*?(الاعضاء|الأعضاء)\s+\S+(\s*[-،]\s*\S+)*", "", s)
    return s.strip()


def residual_len(s: str) -> int:
    c = clean_boiler(s)
    c = re.sub(r"ملف الحكم|لعرض الملف|اضغط هنا|تشريعات مرتبطة", "", c)
    c = re.sub(r"مادة رقم \d+", "", c)
    for law in [
        "قانون العقوبات", "قانون المخدرات والمؤثرات العقلية والسلائف", "قانون العمل",
        "قانون الموجبات والعقود", "قانون الضمان الاجتماعي", "اصول المحاكمات الجزائية",
        "قانون اصول المحاكمات المدنية", "الاسلحة والذخائر", "قانون التجارة",
        "حماية الاحداث المخالفين للقانون او المعرضين للخطر",
        "الدخول الى لبنان والاقامة فيه والخروج منه", "مزاولة مهنة الصيدلة",
        "قانون السير الجديد", "قانون عقود العمل الجماعية", "المعاملات الالكترونية",
    ]:
        c = c.replace(law, "")
    c = re.sub(r"[•·\-،.\s\d]", "", c)
    return len(c)


def load_filtered(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    H = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[H["is_complete"]]:
            continue
        if not r[H["has_article"]]:
            continue
        if residual_len(r[H["page_summary"]]) < 15:
            continue
        out.append({hdr[i]: r[i] for i in range(len(hdr))})
    return out, hdr


def segment_one(client, summary: str):
    """One API call. Returns (dict, usage) or raises."""
    msg = client.messages.create(
        model=MODEL,
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": clean_boiler(summary)}],
    )
    text = "".join(b.text for b in msg.content if b.type == "text").strip()
    # strip code fences if the model added them
    text = re.sub(r"^```(?:json)?|```$", "", text, flags=re.MULTILINE).strip()
    data = json.loads(text)
    return data, msg.usage


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", type=Path, required=True)
    ap.add_argument("--output", type=Path, default=Path("jinayat_segmented.xlsx"))
    ap.add_argument("--cache", type=Path, default=Path("segment_cache.jsonl"))
    ap.add_argument("--limit", type=int, default=None, help="process only first N (test)")
    ap.add_argument("--api-key", type=str, default=None)
    ap.add_argument("--delay", type=float, default=0.3)
    args = ap.parse_args()

    key = args.api_key or os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        sys.exit("No API key. Set $env:ANTHROPIC_API_KEY or pass --api-key.")

    try:
        import anthropic
    except ImportError:
        sys.exit("pip install anthropic")

    client = anthropic.Anthropic(api_key=key)

    rows, hdr = load_filtered(args.input)
    if args.limit:
        rows = rows[: args.limit]
    print(f"Loaded {len(rows)} filtered jinayat rulings"
          + (f" (LIMITED to {args.limit} for testing)" if args.limit else ""))

    # resume from cache
    done = {}
    if args.cache.exists():
        for line in args.cache.read_text(encoding="utf-8").splitlines():
            try:
                rec = json.loads(line)
                done[str(rec["ruling_id"])] = rec
            except Exception:
                pass
    if done:
        print(f"Resuming: {len(done)} already segmented (cached, not re-charged)")

    tot_in = tot_out = 0
    status_counts = {"clean": 0, "partial": 0, "fused": 0, "error": 0}

    with open(args.cache, "a", encoding="utf-8") as cache_f:
        for i, row in enumerate(rows, 1):
            rid = str(row["ruling_id"])
            if rid in done:
                status_counts[done[rid].get("status", "error")] = \
                    status_counts.get(done[rid].get("status", "error"), 0) + 1
                continue
            try:
                data, usage = segment_one(client, row["page_summary"])
                tot_in += usage.input_tokens
                tot_out += usage.output_tokens
                rec = {
                    "ruling_id": row["ruling_id"],
                    "situation": data.get("situation", ""),
                    "reasoning_verdict": data.get("reasoning_verdict", ""),
                    "status": data.get("status", "partial"),
                }
            except Exception as e:
                rec = {"ruling_id": row["ruling_id"], "situation": "",
                       "reasoning_verdict": "", "status": "error", "error": str(e)}
            cache_f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            cache_f.flush()
            done[rid] = rec
            status_counts[rec["status"]] = status_counts.get(rec["status"], 0) + 1

            if i % 10 == 0 or i == len(rows):
                est = tot_in / 1e6 * PRICE_IN_PER_MTOK + tot_out / 1e6 * PRICE_OUT_PER_MTOK
                print(f"  [{i}/{len(rows)}] in={tot_in} out={tot_out} tok "
                      f"~${est:.3f}  status={status_counts}")
            time.sleep(args.delay)

    # ---- write the merged Excel ----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Jinayat Segmented"
    ws.sheet_view.rightToLeft = True
    new_hdr = hdr + ["situation", "reasoning_verdict", "segment_status"]
    ws.append(new_hdr)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="DDDDDD", end_color="DDDDDD")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        rid = str(row["ruling_id"])
        seg = done.get(rid, {})
        ws.append([row[h] for h in hdr]
                  + [seg.get("situation", ""), seg.get("reasoning_verdict", ""),
                     seg.get("status", "error")])
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ["S", "T", "U", "V"]:
        ws.column_dimensions[col].width = 60
    wb.save(args.output)

    est = tot_in / 1e6 * PRICE_IN_PER_MTOK + tot_out / 1e6 * PRICE_OUT_PER_MTOK
    print(f"\nDONE. Wrote {args.output}")
    print(f"  status breakdown: {status_counts}")
    print(f"  tokens this run: in={tot_in} out={tot_out}  est cost ~${est:.3f}")
    print(f"  (cache: {args.cache} — delete it to force a full re-run)")


if __name__ == "__main__":
    main()